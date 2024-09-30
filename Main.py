import os
from openpyxl import Workbook
import pdfplumber
from datetime import datetime
from tkinter import Tk, Button, Label, filedialog, messagebox

# Função para processar os arquivos PDF
def process_files(pdf_files):
    try:
        if not pdf_files:
            raise Exception("Nenhum arquivo selecionado")

        # Criando arquivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'PDF Data'

        # Cabeçalhos no Excel
        ws['B1'] = 'Código'
        ws['C1'] = 'Descrição'

        last_empty_line = 2  # Linha inicial para dados

        # Processa cada arquivo PDF
        for file in pdf_files:
            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:  # Iterar por todas as páginas
                    tables = page.extract_tables()  # Extrai as tabelas da página

                    # Verifica se há tabelas na página
                    if tables:
                        for table in tables:
                            # Percorre as linhas da tabela
                            for row in table:
                                if len(row) >= 2:  # Verifica se há pelo menos duas colunas (Código e Descrição)
                                    codigo = row[0].strip()  # Primeira coluna é o Código
                                    descricao = row[1].strip()  # Segunda coluna é a Descrição

                                    # Escreve no Excel
                                    ws[f'A{last_empty_line}'] = os.path.basename(file)
                                    ws[f'B{last_empty_line}'] = codigo
                                    ws[f'C{last_empty_line}'] = descricao
                                    last_empty_line += 1

        # Pergunta ao usuário onde salvar o arquivo Excel
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Salvar arquivo Excel"
        )
        
        if save_path:  # Se o usuário escolheu um local
            wb.save(save_path)
            messagebox.showinfo("Sucesso", f"Arquivo Excel salvo com sucesso em {save_path}!")
        else:
            messagebox.showwarning("Cancelado", "Salvamento cancelado pelo usuário.")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar os arquivos: {e}")

# Função para selecionar os arquivos PDF e processá-los
def select_files_and_process():
    pdf_files = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if pdf_files:
        process_files(pdf_files)

# Interface gráfica
def create_interface():
    root = Tk()
    root.title("Processar Dados de PDF")

    Label(root, text="Clique no botão para selecionar os arquivos PDF e processá-los:").pack(pady=10)

    Button(root, text="Selecionar arquivos e processar", command=select_files_and_process).pack(pady=20)

    root.geometry("500x200")
    root.mainloop()

if __name__ == "__main__":
    create_interface()
